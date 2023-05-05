from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import openpyxl

#Select your excel file
book = openpyxl.load_workbook('scrap.xlsx')
sheet = book['Sheet1']

#Enter your installed chromedriver path
service = Service("/Users/youngman/Desktop/dlg/chromedriver")
driver = webdriver.Chrome(service=service)
n = 0
column_title_name = 1
column_title_twitter = 2
column_title_instagram = 3
row = 0
column = 1
judge = True
exist_judge = False
sheet.cell(row=row+1, column=column_title_name, value='name')
sheet.cell(row=row+1, column=column_title_twitter, value='twitter')
sheet.cell(row=row+1, column=column_title_instagram, value='instagram')
for _ in range(30):
  n += 1
  #Enter your Website
  driver_path = f'https://www.waseda.jp/inst/weekly/circleguide/list/?pg={n}'
  driver.get(driver_path)
  driver.implicitly_wait(10)
  elements = '.l-weekly-searchResults-list > .col-sm-3 > .eq > .list-image'
  count = driver.find_elements(By.CSS_SELECTOR, elements)
  print('count ' + str(len(count)))

#Existed 24 sector in the Website
  for i in range(24):
    row += 1
    time.sleep(1)
    counter = f'{n}-{i}'
    print(counter)
    count = driver.find_elements(By.CSS_SELECTOR, elements)
    driver.implicitly_wait(10)
    count[i].click()
    driver.implicitly_wait(10)
    id_elements = 'anc_6'
    id_exist = driver.find_elements(By.ID, id_elements)
    link = '.bd-base > .list-with-arrow > .push-half-bottom > .link-about'
    circle_title = '.wrap-outer > .wrap-inner > .soft-double-top  > .container > .row > .col-sm-12 > .m-circleDetail > .searchResults-title > .ja'
    try:
      link_result_0 = driver.find_elements(By.CSS_SELECTOR, link)[0].get_attribute("href")
      circle_title_result = driver.find_elements(By.CSS_SELECTOR, circle_title)[0].text
      print(circle_title_result)
      print(link_result_0)
      judge = True
    except IndexError:
      print('Link Null')
      judge = False
    else: 
      sheet.cell(row=row+1, column=column, value=circle_title_result)
      exist_judge == False
      if 'twitter' in link_result_0:
        sheet.cell(row=row+1, column=column+1, value=link_result_0)
      elif 'www.instagram.com' in link_result_0:
        sheet.cell(row=row+1, column=column+2, value=link_result_0)
      else: 
        sheet.cell(row=row+1, column=column+3, value=link_result_0) 
        exist_judge = True

    if judge == True:
      try:
        link_result_1 = driver.find_elements(By.CSS_SELECTOR, link)[1].get_attribute('href')
        if 'twitter' in link_result_1:
          sheet.cell(row=row+1, column=column+1, value=link_result_1)
        elif 'www.instagram.com' in link_result_1:
          sheet.cell(row=row+1, column=column+2, value=link_result_1)
        elif exist_judge == False:
          sheet.cell(row=row+1, column=column+3, value=link_result_1) 
        else:
          sheet.cell(row=row+1, column=column+4, value=link_result_1) 
        print(link_result_1)
      except IndexError:
        print('')

      try:
        link_result_2 = driver.find_elements(By.CSS_SELECTOR, link)[2].get_attribute('href')
        if 'twitter' in link_result_2:
          sheet.cell(row=row+1, column=column+1, value=link_result_2)
        elif 'www.instagram.com' in link_result_2:
          sheet.cell(row=row+1, column=column+2, value=link_result_2)
        elif exist_judge == False:
          sheet.cell(row=row+1, column=column+3, value=link_result_2) 
        else:
          sheet.cell(row=row+1, column=column+4, value=link_result_2) 
        print(link_result_2)
      except IndexError:
        print('')

      try:
        link_result_3 = driver.find_elements(By.CSS_SELECTOR, link)[3].get_attribute('href')
        if 'twitter' in link_result_3:
          sheet.cell(row=row+1, column=column+1, value=link_result_3)
        elif 'www.instagram.com' in link_result_3:
          sheet.cell(row=row+1, column=column+2, value=link_result_3)
        elif exist_judge == False:
          sheet.cell(row=row+1, column=column+3, value=link_result_3) 
        else:
          sheet.cell(row=row+1, column=column+4, value=link_result_3) 
        print(link_result_3)
      except IndexError:
        print('')

    driver.get(driver_path)
    driver.implicitly_wait(10)
    book.save('scrap.xlsx')